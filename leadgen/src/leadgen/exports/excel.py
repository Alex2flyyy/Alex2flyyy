"""Excel export.

Produces a workbook you can hand to someone who has never seen this system:
a formatted Leads sheet with frozen headers, autofilter, score color-coding,
and clickable links, plus a Summary sheet with the breakdowns.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from leadgen.logging import get_logger

log = get_logger(__name__)

HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
LINK_FONT = Font(color="2563EB", underline="single")

SCORE_FILLS = [
    (85, PatternFill("solid", start_color="FECACA")),  # hot
    (70, PatternFill("solid", start_color="FED7AA")),  # warm
    (55, PatternFill("solid", start_color="D1FAE5")),  # qualified
]

COLUMN_WIDTHS = {
    "rank": 6, "lead_score": 10, "business_name": 34, "owner_name": 18,
    "phone": 16, "email": 30, "website": 34, "website_status": 17,
    "website_score": 12, "why_they_need_a_website": 60, "street_address": 28,
    "city": 16, "state": 7, "zip_code": 9, "county": 16, "niche": 16,
    "category": 24, "google_rating": 9, "review_count": 9,
    "google_maps_url": 26, "hours": 34, "social_links": 30,
    "business_summary": 60, "top_problems": 60,
    "redesign_recommendations": 60, "outreach_angle": 60,
    "outreach_subject": 34, "talking_points": 60,
    "estimated_value_tier": 12, "lead_status": 14,
}


def write_excel(
    rows: list[dict[str, Any]],
    path: Path | str,
    *,
    stats: dict[str, Any] | None = None,
    columns: list[str] | None = None,
) -> Path:
    from leadgen.reports.daily_report import EXPORT_COLUMNS

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = columns or EXPORT_COLUMNS

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    _write_leads(sheet, rows, columns)

    if stats:
        _write_summary(workbook.create_sheet("Summary"), stats, len(rows))

    workbook.save(path)
    log.info("export.excel", path=str(path), rows=len(rows))
    return path


def _write_leads(sheet: Worksheet, rows: list[dict[str, Any]], columns: list[str]) -> None:
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=column.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(column, 16)

    sheet.freeze_panes = "C2"
    sheet.row_dimensions[1].height = 28

    for row_index, row in enumerate(rows, start=2):
        for col_index, column in enumerate(columns, start=1):
            value = row.get(column, "")
            if isinstance(value, list | tuple):
                value = " | ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = "; ".join(f"{k}: {v}" for k, v in value.items())

            cell = sheet.cell(row=row_index, column=col_index, value=value)

            if column in {"website", "google_maps_url"} and value:
                cell.hyperlink = str(value)
                cell.font = LINK_FONT
            elif column == "email" and value:
                cell.hyperlink = f"mailto:{value}"
                cell.font = LINK_FONT
            elif column == "lead_score":
                score = row.get("lead_score") or 0
                for threshold, fill in SCORE_FILLS:
                    if score >= threshold:
                        cell.fill = fill
                        break
                cell.font = Font(bold=True)
            elif column in {"why_they_need_a_website", "business_summary", "top_problems",
                            "outreach_angle", "talking_points", "redesign_recommendations"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"


def _write_summary(sheet: Worksheet, stats: dict[str, Any], lead_count: int) -> None:
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 14

    def heading(row: int, text: str) -> int:
        cell = sheet.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12)
        return row + 1

    row = heading(1, f"Lead report — {datetime.utcnow():%Y-%m-%d}")
    row += 1

    overview = stats.get("overview", {})
    row = heading(row, "Overview")
    for label, key in [
        ("Leads in this export", None),
        ("Total businesses", "total_businesses"),
        ("Total leads", "total_leads"),
        ("Qualified leads", "qualified_leads"),
        ("New today", "new_today"),
        ("Contacted", "contacted"),
        ("Pending contact", "pending"),
        ("Won", "won"),
        ("Average score", "avg_score"),
        ("Contact rate %", "contact_rate"),
        ("Win rate %", "win_rate"),
    ]:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=lead_count if key is None else overview.get(key, 0))
        row += 1

    for title, key, fields in [
        ("Leads by city", "by_city", ("city", "leads", "avg_score")),
        ("Leads by niche", "by_niche", ("niche", "leads", "avg_score")),
        ("Website status", "by_website_status", ("display", "count", None)),
    ]:
        row += 1
        row = heading(row, title)
        for entry in stats.get(key, []):
            sheet.cell(row=row, column=1, value=entry.get(fields[0], ""))
            sheet.cell(row=row, column=2, value=entry.get(fields[1], 0))
            if fields[2]:
                sheet.cell(row=row, column=3, value=entry.get(fields[2], 0))
            row += 1
